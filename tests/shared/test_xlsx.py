
import json
import openpyxl
from openpyxl import load_workbook

import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def convert_xlsx_to_json(filepath, include_style=False, sheet_names=None):
    """
    将Excel转换为适合LLM的JSON

    Args:
        filepath: Excel文件路径
        include_style: 是否包含样式信息
        sheet_names: 指定要转换的工作表名称列表，None表示所有工作表
    """
    try:
        wb = load_workbook(filepath, data_only=True)
        result = {
            "filename": filepath,
            "sheet_count": len(wb.sheetnames),
            "sheets": []
        }

        # 确定要处理的工作表
        if sheet_names is None:
            sheets_to_process = wb.sheetnames
        else:
            sheets_to_process = [name for name in sheet_names if name in wb.sheetnames]

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]

            # 跳过完全为空的工作表
            if ws.max_row == 0 and ws.max_column == 0:
                continue

            # 获取最大列的字母表示
            max_column_letter = get_column_letter(ws.max_column) if ws.max_column > 0 else "A"

            sheet_info = {
                "name": ws.title,
                "index": wb.sheetnames.index(sheet_name) + 1,
                "dimensions": f"A1:{max_column_letter}{ws.max_row}" if ws.max_row > 0 else "A1:A1",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "max_column_letter": max_column_letter,
                "merged_cells_count": len(ws.merged_cells.ranges),
                "merged_ranges": [str(mr) for mr in ws.merged_cells.ranges],
                "data": []
            }

            # 以列表形式提取数据，包含空单元格
            data = []
            for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                  min_col=1, max_col=ws.max_column):
                row = []
                for cell in r:
                    cell_data = {
                        "address": cell.coordinate,
                        "value": cell.value,
                        "column": get_column_letter(cell.column),
                        "row": cell.row,
                        "column_index": cell.column,
                        "is_merged": cell.coordinate in ws.merged_cells
                    }

                    # 检查是否是公式（虽然data_only=True会返回计算结果）
                    if cell.data_type == 'f':
                        cell_data["has_formula"] = True
                    else:
                        cell_data["has_formula"] = False

                    if include_style:
                        cell_data["style"] = {
                            "font_size": cell.font.sz or 11,
                            "font_name": cell.font.name,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "number_format": cell.number_format,
                            "alignment_horizontal": cell.alignment.horizontal,
                            "alignment_vertical": cell.alignment.vertical
                        }
                    row.append(cell_data)
                data.append(row)

            sheet_info["data"] = data
            result["sheets"].append(sheet_info)

        return json.dumps(result, indent=2, default=str, ensure_ascii=False)

    except FileNotFoundError:
        return json.dumps({"error": f"文件不存在: {filepath}"})
    except Exception as e:
        return json.dumps({"error": f"处理文件时出错: {str(e)}"})


def save_json_to_file(json_str, output_path):
    """将JSON字符串保存到文件"""
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(json_str)
        print(f"JSON已保存到: {output_path}")
        return True
    except Exception as e:
        print(f"保存JSON文件时出错: {str(e)}")
        return False


if __name__ == "__main__":
    my_excel_file = "/home/rd/Downloads/2.xlsx"

    # 转换为JSON
    json_data = convert_xlsx_to_json(my_excel_file, include_style=False)

    # 打印前1000个字符（避免控制台输出过长）
    print("JSON预览（前1000字符）:")
    print(json_data[:1000])
    if len(json_data) > 1000:
        print(f"... 还有 {len(json_data) - 1000} 字符")

    # 保存到文件
    output_file = my_excel_file.replace('.xlsx', '.json').replace('.xls', '.json')
    save_json_to_file(json_data, output_file)

    # 解析并显示基本信息
    data_obj = json.loads(json_data)
    if "error" not in data_obj:
        print(f"\nExcel文件基本信息:")
        print(f"- 文件: {data_obj.get('filename', '未知')}")
        print(f"- 工作表数量: {data_obj.get('sheet_count', 0)}")
        for sheet in data_obj.get("sheets", []):
            print(f"  - {sheet['name']}: {sheet['dimensions']}, 合并单元格: {sheet['merged_cells_count']}个")