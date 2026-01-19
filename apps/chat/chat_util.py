#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) [2025] [liuyngchng@hotmail.com] - All rights reserved.
import json
import os
import logging.config
import sys
from typing import Generator

import requests

from common.docx_md_util import convert_docx_to_md, get_md_file_content

ALLOWED_EXTENSIONS = {
    'txt', 'md', 'py', 'js', 'html', 'css', 'json',
    'pdf', 'xlsx', 'docx',
    'jpg', 'jpeg', 'png', 'gif'
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# 配置日志
log_config_path = 'logging.conf'
if os.path.exists(log_config_path):
    logging.config.fileConfig(log_config_path, encoding="utf-8", disable_existing_loggers=False)
    print(f"使用日志配置文件: {log_config_path}")
else:
    print("日志配置文件不存在，使用默认配置")
    from common.const import LOG_FORMATTER
    logging.basicConfig(level=logging.INFO,format=LOG_FORMATTER,force=True,stream=sys.stdout)


logger = logging.getLogger(__name__)

# ============== 配置常量 ==============
# LLM API 配置（可替换为任何兼容OpenAI API的接口）
class LLMConfig:

    # 请求参数配置
    MAX_TOKENS = int(os.getenv("MAX_TOKENS", 8000))
    TEMPERATURE = float(os.getenv("TEMPERATURE", 0.7))
    TOP_P = float(os.getenv("TOP_P", 0.9))

    # 流式响应配置
    STREAM = True
    TIMEOUT = 600  # 请求超时时间（秒）

    # 系统提示词
    SYSTEM_PROMPT = os.getenv("SYSTEM_PROMPT", "你是一个有用的AI助手。请用中文回答用户的问题。")


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_text_from_file(filepath, filename):
    """从文件中提取文本内容"""
    try:
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        logger.info(f"正在提取文件: {filename}, 扩展名: {ext}")

        # 文本文件直接读取
        if ext in ['txt', 'md', 'py', 'js', 'html', 'css', 'json']:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                logger.info(f"文本文件读取成功，长度: {len(content)} 字符")
                return content

        # PDF文件（需要安装PyPDF2或pdfplumber）
        elif ext == 'pdf':
            try:
                import PyPDF2
                text = []
                with open(filepath, 'rb') as f:
                    pdf = PyPDF2.PdfReader(f)
                    for page_num, page in enumerate(pdf.pages):
                        page_text = page.extract_text()
                        text.append(page_text)
                result = '\n'.join(text)
                logger.info(f"PDF文件解析成功，页数: {len(pdf.pages)}，长度: {len(result)} 字符")
                return result
            except ImportError:
                logger.warning("需要安装PyPDF2库来解析PDF文件")
                return "[需要安装PyPDF2库来解析PDF文件]"

        # Word文档（需要安装python-docx）
        elif ext in ['docx']:
            try:
                md_file = convert_docx_to_md(filepath, True)
                result = get_md_file_content(md_file)
                return result
            except ImportError:
                logger.warning("需要安装 pypandoc 库来解析Word文档")
                return "[需要安装 pypandoc 库来解析Word文档]"

        # Excel文件（需要安装openpyxl）
        elif ext in ['xlsx']:
            return get_xlsx_content(filepath)

        # 图片文件（需要安装PIL和pytesseract）
        elif ext in ['jpg', 'jpeg', 'png', 'gif']:
            try:
                from PIL import Image
                import pytesseract
                image = Image.open(filepath)
                text = pytesseract.image_to_string(image, lang='chi_sim+eng')
                logger.info(f"图片文件识别成功，识别到 {len(text)} 字符")
                return text if text else "[图片中未识别到文字]"
            except ImportError:
                logger.warning("需要安装PIL和pytesseract库来识别图片文字")
                return "[需要安装PIL和pytesseract库来识别图片文字]"

        else:
            logger.warning(f"不支持的文件格式: {ext}")
            return f"[不支持的文件格式: {ext}]"

    except Exception as e:
        logger.error(f"读取文件时出错: {str(e)}")
        return f"[读取文件时出错: {str(e)}]"


def get_xlsx_content(filepath) -> str:
    """
    获取 xlsx 文件的内容
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=filepath, read_only=True, data_only=True)
        text_lines = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_lines.append(f"## {sheet_name}")

            # 获取最大列宽
            max_col = ws.max_column
            max_row = ws.max_row

            if max_row > 0 and max_col > 0:
                # 创建Markdown表格 - 修复这里
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=max_row,
                                        min_col=1, max_col=max_col,
                                        values_only=True):
                    # 不要转义管道符，而是处理空值和特殊字符
                    formatted_row = []
                    for cell in row:
                        if cell is None:
                            formatted_row.append("")
                        else:
                            # 将单元格内容转换为字符串，并处理换行
                            cell_str = str(cell)
                            # 替换换行符为空格，避免破坏表格格式
                            cell_str = cell_str.replace('\n', ' ').replace('\r', ' ')
                            # 清理多余的空白字符
                            cell_str = ' '.join(cell_str.split())
                            formatted_row.append(cell_str)
                    rows.append(formatted_row)

                # 生成Markdown表格
                if rows:
                    # 表头
                    md_lines = ['| ' + ' | '.join(rows[0]) + ' |']
                    # 表头分隔线
                    header_separator = ['---'] * len(rows[0])
                    md_lines.append('| ' + ' | '.join(header_separator) + ' |')
                    # 数据行
                    for row in rows[1:]:
                        md_lines.append('| ' + ' | '.join(row) + ' |')

                    text_lines.append('\n'.join(md_lines))
                text_lines.append('')  # 空行分隔

        result = '\n'.join(text_lines)
        logger.info(f"Excel文件解析成功，工作表数: {len(wb.sheetnames)}，长度: {len(result)} 字符")
        return result
    except ImportError:
        logger.warning("需要安装openpyxl库来解析Excel文件")
        return "[需要安装openpyxl库来解析Excel文件]"
    except Exception as e:
        logger.error(f"读取Excel文件时出错: {str(e)}")
        return f"[读取Excel文件时出错: {str(e)}]"


def generate_stream_response(messages: list, llm_cfg: dict, max_tokens: int = None) -> Generator[str, None, None]:
    """
    生成流式响应
    """
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {llm_cfg['llm_api_key']}"
    }

    payload = {
        "model": llm_cfg['llm_model_name'],
        "messages": messages,
        "max_tokens": max_tokens or LLMConfig.MAX_TOKENS,
        "temperature": LLMConfig.TEMPERATURE,
        "top_p": LLMConfig.TOP_P,
        "stream": LLMConfig.STREAM,
    }

    try:
        logger.info(f"向LLM API发送请求，模型: {llm_cfg['llm_model_name']}, 消息数量: {len(messages)}")
        logger.info(f"请求参数: max_tokens={payload['max_tokens']}, temperature={LLMConfig.TEMPERATURE}")

        response = requests.post(
            f"{llm_cfg['llm_api_uri']}/chat/completions",
            headers=headers,
            json=payload,
            stream=True,
            verify=False,
            timeout=LLMConfig.TIMEOUT
        )

        # 记录响应状态码
        logger.info(f"llm_api_response_code: {response.status_code}")

        # 如果响应不是200，记录详细错误信息
        if response.status_code != 200:
            error_content = response.text
            logger.error(f"llm_api_response_error, {response.status_code}: {error_content}")

            try:
                error_json = response.json()
                logger.error(f"llm_api_response_error_json: {json.dumps(error_json, indent=2, ensure_ascii=False)}")

                # 提取具体错误信息
                if 'error' in error_json:
                    error_msg = error_json['error']
                    if isinstance(error_msg, dict) and 'message' in error_msg:
                        error_detail = error_msg['message']
                    elif isinstance(error_msg, str):
                        error_detail = error_msg
                    else:
                        error_detail = str(error_msg)

                    logger.error(f"错误消息: {error_detail}")
                else:
                    logger.error(f"完整错误响应: {json.dumps(error_json, ensure_ascii=False)}")

            except json.JSONDecodeError:
                logger.error(f"错误响应不是JSON格式，原始内容: {error_content}")
            except Exception as e:
                logger.error(f"解析错误响应时出错: {str(e)}, 原始内容: {error_content}")

            # 向客户端返回错误信息
            error_message = f"API请求失败: {response.status_code} - {response.reason}"
            yield f"data: {json.dumps({'error': error_message})}\n\n"
            yield "data: [DONE]\n\n"
            return

        response.raise_for_status()

        for line in response.iter_lines():
            if line:
                line = line.decode('utf-8')
                if line.startswith('data: '):
                    data = line[6:]  # 移除 'data: ' 前缀
                    if data != '[DONE]':
                        try:
                            chunk = json.loads(data)
                            if 'choices' in chunk and len(chunk['choices']) > 0:
                                delta = chunk['choices'][0].get('delta', {})
                                if 'content' in delta and delta['content']:
                                    yield f"data: {json.dumps({'content': delta['content']})}\n\n"
                        except json.JSONDecodeError:
                            logger.warning("解析JSON数据失败")
                            continue

        yield "data: [DONE]\n\n"

    except requests.exceptions.RequestException as e:
        error_msg = f"API请求错误: {str(e)}"
        logger.error(error_msg)

        # 如果有响应对象，记录更多信息
        if hasattr(e, 'response') and e.response is not None:
            try:
                error_json = e.response.json()
                logger.error(f"请求异常中的错误详情: {json.dumps(error_json, ensure_ascii=False)}")
            except:
                logger.error(f"请求异常中的响应内容: {e.response.text}")

        yield f"data: {json.dumps({'error': error_msg})}\n\n"
        yield "data: [DONE]\n\n"
