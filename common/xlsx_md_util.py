#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) [2025] [liuyngchng@hotmail.com] - All rights reserved.
import logging.config
import os
from pathlib import Path

logging.config.fileConfig('logging.conf', encoding="utf-8")
logger = logging.getLogger(__name__)

OUTPUT_DIR = "output_doc"

def convert_xlsx_to_md(excel_path: str, include_sheet_names: bool = True, output_abs_path: bool = False) -> str:
    """
    将 Excel 中的多个 sheet 转换为 markdown 格式的文本并保存到文件
    能够更好地处理合并单元格
    :param excel_path: Excel 文件路径
    :param include_sheet_names: 是否包含工作表名称
    :return: markdown 文件的磁盘路径
    """
    import pandas as pd
    try:
        # 确保输出目录存在
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # 获取原文件名（不含扩展名）并生成 markdown 文件名
        excel_file = Path(excel_path)
        md_filename = excel_file.stem + ".md"
        md_path = os.path.join(OUTPUT_DIR, md_filename)

        # 读取所有工作表
        excel_file_obj = pd.ExcelFile(excel_path)
        markdown_parts = []

        for sheet_name in excel_file_obj.sheet_names:
            # 使用 openpyxl 引擎来读取合并单元格信息
            df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')

            if include_sheet_names:
                markdown_parts.append(f"## 工作表: {sheet_name}")
                markdown_parts.append("")

            if not df.empty:
                # 处理 NaN 值
                df = df.fillna('')

                # 使用改进的Markdown表格生成函数
                markdown_table = dataframe_to_markdown_with_merged_cells(df, excel_path, sheet_name)
                markdown_parts.append(markdown_table)
                markdown_parts.append("")  # 空行分隔

        # 将内容写入文件 合并的单元格会产生 Unnamed:
        markdown_content = "\n".join(markdown_parts).replace("Unnamed:", "")
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        abs_path = os.path.abspath(md_path)
        logger.info(f"成功转换 Excel 文件: {excel_path} -> {abs_path}, 包含 {len(excel_file_obj.sheet_names)} 个工作表")
        if output_abs_path:
            return abs_path
        else:
            return md_path

    except Exception as e:
        logger.error(f"excel_to_md_error, file {excel_path}, {str(e)}")
        return ""


def dataframe_to_markdown_with_merged_cells(df, excel_path, sheet_name):
    """
    将DataFrame转换为Markdown表格，尝试处理合并单元格
    """
    try:
        import openpyxl

        # 使用openpyxl直接读取Excel文件来获取合并单元格信息
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook[sheet_name]

        # 获取合并单元格范围
        merged_ranges = worksheet.merged_cells.ranges

        # 创建合并单元格映射
        merged_cells_map = {}
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            # 注意：openpyxl的行列索引从1开始，pandas从0开始
            top_left_value = worksheet.cell(min_row, min_col).value

            # 记录合并区域内的所有单元格（除了左上角）
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    if row == min_row and col == min_col:
                        continue  # 跳过左上角单元格
                    merged_cells_map[(row - 1, col - 1)] = top_left_value  # 转换为0-based索引

        # 生成Markdown表格
        markdown_lines = []

        # 表头
        headers = df.columns.tolist()
        header_row = "| " + " | ".join(str(h) for h in headers) + " |"
        markdown_lines.append(header_row)

        # 分隔线
        separator = "|" + "|".join([" --- " for _ in headers]) + "|"
        markdown_lines.append(separator)

        # 数据行
        for row_idx, (index, row) in enumerate(df.iterrows()):
            row_data = []
            for col_idx, value in enumerate(row):
                cell_key = (row_idx, col_idx)
                if cell_key in merged_cells_map:
                    # 如果是合并单元格且不是左上角，留空
                    row_data.append("")
                else:
                    row_data.append(str(value) if pd.notna(value) else "")

            row_str = "| " + " | ".join(row_data) + " |"
            markdown_lines.append(row_str)

        return "\n".join(markdown_lines)

    except Exception as e:
        logger.warning(f"处理合并单元格时出错，使用标准Markdown转换: {str(e)}")
        # 回退到标准的Markdown转换
        return df.to_markdown(index=False)


def convert_xlsx_to_md_simple(excel_path: str, include_sheet_names: bool = True, output_abs_path: bool = False) -> str:
    """
    简化版本：使用标准Markdown转换，但在表格前后添加说明
    """
    import pandas as pd
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        excel_file = Path(excel_path)
        md_filename = excel_file.stem + ".md"
        md_path = os.path.join(OUTPUT_DIR, md_filename)

        excel_file_obj = pd.ExcelFile(excel_path)
        markdown_parts = []

        for sheet_name in excel_file_obj.sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)

            if include_sheet_names:
                markdown_parts.append(f"## 工作表: {sheet_name}")
                markdown_parts.append("")

            if not df.empty:
                df = df.fillna('')

                # 添加表格说明
                markdown_parts.append("> 💡 **注意**: 原始Excel中的合并单元格在Markdown中可能显示异常")
                markdown_parts.append("")

                markdown_table = df.to_markdown(index=False)
                markdown_parts.append(markdown_table)
                markdown_parts.append("")  # 空行分隔

        markdown_content = "\n".join(markdown_parts).replace("Unnamed:", "")
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        abs_path = os.path.abspath(md_path)
        logger.info(f"成功转换 Excel 文件: {excel_path} -> {abs_path}")
        if output_abs_path:
            return abs_path
        else:
            return md_path

    except Exception as e:
        logger.error(f"excel_to_md_error, file {excel_path}, {str(e)}")
        return ""


def convert_xlsx_to_md_advanced(excel_path: str, include_sheet_names: bool = True,
                                output_abs_path: bool = False) -> str:
    """
    高级版本：将Excel转换为更易读的Markdown格式
    """
    import pandas as pd
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        excel_file = Path(excel_path)
        md_filename = excel_file.stem + ".md"
        md_path = os.path.join(OUTPUT_DIR, md_filename)

        excel_file_obj = pd.ExcelFile(excel_path)
        markdown_parts = []

        for sheet_name in excel_file_obj.sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')

            if include_sheet_names:
                markdown_parts.append(f"## 📊 工作表: {sheet_name}")
                markdown_parts.append("")

            if not df.empty:
                df = df.fillna('')

                # 检查数据维度
                if df.shape[1] <= 3:  # 列数较少时使用更好的格式
                    markdown_parts.extend(dataframe_to_readable_list(df, sheet_name))
                else:
                    markdown_parts.append("> ⚠️ **表格预览** (复杂表格建议查看原文件)")
                    markdown_parts.append("")
                    markdown_table = df.to_markdown(index=False)
                    markdown_parts.append(markdown_table)

                markdown_parts.append("")  # 空行分隔

        markdown_content = "\n".join(markdown_parts).replace("Unnamed:", "")
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(markdown_content)
        abs_path = os.path.abspath(md_path)
        logger.info(f"成功转换 Excel 文件: {excel_path} -> {abs_path}")
        return abs_path if output_abs_path else md_path

    except Exception as e:
        logger.error(f"excel_to_md_error, file {excel_path}, {str(e)}")
        return ""


def dataframe_to_readable_list(df, sheet_name):
    """将DataFrame转换为更易读的列表格式"""
    import pandas as pd
    lines = [f"### {sheet_name} 数据", ""]
    headers = df.columns.tolist()
    for index, row in df.iterrows():
        lines.append(f"**记录 {index + 1}:**")
        for header, value in zip(headers, row):
            if pd.notna(value) and str(value).strip():
                lines.append(f"- **{header}**: {value}")
        lines.append("")

    return lines

# 使用示例
if __name__ == "__main__":
    my_excel_file = "/home/rd/Downloads/1.xlsx"  # 替换为你的 Excel 文件路径
    md_file_path = convert_xlsx_to_md_advanced(my_excel_file, True)
    if md_file_path:
        logger.info(f"Markdown文件已保存到: {md_file_path}")

        # 可选：读取并显示部分内容
        with open(md_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            logger.info(f"文件前500字符预览:\n{content[:500]}...")
    else:
        logger.info("转换失败")